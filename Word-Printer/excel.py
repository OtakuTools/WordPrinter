from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Font

class excel:
    def __init__(self):
        self.colorDict = {
            'FFFFFF00':'title',
            'FFD40000':'BasicInfo.PartyA.projectName',
            #'FFD30000':'BasicInfo.PartyA.company',
            'FFC90000':'BasicInfo.Detail.amount',
            'FFBF0000':'BasicInfo.Team.PM',
            'FFBD0000':'ServiceProcess.Report.time',
            'FFA90000':'ServiceProcess.Config.preReleaseDate',
            'FFA80000':'ServiceProcess.Config.applicationDate',
            'FFA70000':'ServiceProcess.Config.SN',
            }

    def title( self, title ):
        for worksheet in self.xlsx.worksheets:
            if worksheet['A2'].fill.fgColor.rgb == 'FFFFFF00':
                worksheet['A2'] = title
                worksheet['A2'].fill = PatternFill( fill_type=None )
                worksheet['A2'].font = Font( name=worksheet['A2'].font.name ,color=None )

    def replace( self , project ):
        for worksheet in self.xlsx.worksheets:
            for row in worksheet.rows:
                for curCell in row:
                    if curCell.fill.fgColor.rgb == 'FFFFFF00':
                        if curCell.font.color.rgb == 'FFFF0000':
                            curCell.value = title
                        elif curCell.font.color.rgb == 'FFD30000':
                            curCell.value = "客户名称：" + str(project.BasicInfo.PartyA.company)
                        else:
                            obj = project
                            for attrName in self.colorDict[str(curCell.font.color.rgb)].split('.'):
                                obj = getattr( obj , attrName, curCell.value )
                            curCell.value = str(obj)

                        curCell.fill = PatternFill( fill_type=None )
                        curCell.font = Font( name=curCell.font.name ,color=None )

    def run( self, src , dst , titleName , project):
        self.xlsx = load_workbook( filename = src )
        self.title( titleName )
        self.replace( project )
        return self.xlsx